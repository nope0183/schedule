import datetime
import hashlib
import json
import os
import re
import sys
import tempfile
from io import BytesIO

import requests
from openpyxl import load_workbook


# ============================================================
# CONFIG
# ============================================================

URL = "https://serp-koll.ru/images/ep/k1/rasp1.xlsx"

JSON_FILE = "schedule.json"
HASH_FILE = "schedule.hash"

# Р”Р»СЏ Р»РѕРєР°Р»СЊРЅРѕР№ РїСЂРѕРІРµСЂРєРё РјРѕР¶РЅРѕ:
#   LOCAL_XLSX=rasp1.xlsx python parse_schedule.py
LOCAL_XLSX = os.getenv("LOCAL_XLSX")

MAX_LESSON = 7
GROUP_RE = re.compile(r"^\d{4}[Рђ-РЇР°-СЏРЃС‘]?$")

TEACHER_RE = re.compile(
    r"^[Рђ-РЇРЃ][Р°-СЏС‘-]+(?:\s+[Рђ-РЇРЃ]\.?\s*[Рђ-РЇРЃ]\.?)$",
    re.IGNORECASE,
)

CODE_ONLY_RE = re.compile(
    r"^[Рђ-РЇРЃ]{1,8}\s*\.?\s*\d{1,3}(?:[.\-]\d{1,3})?[*Р°-СЏРђ-РЇРЃС‘]*$",
    re.IGNORECASE,
)

CODE_WITH_TEXT_RE = re.compile(
    r"^([Рђ-РЇРЃ]{1,8}\s*\.?\s*\d{1,3}(?:[.\-]\d{1,3})?[*Р°-СЏРђ-РЇРЃС‘]*)\s+(.+)$",
    re.IGNORECASE,
)

ROOM_IN_BRACKETS_RE = re.compile(
    r"\(\s*(\d{1,3}[Р°-СЏС‘]?)\s*\)",
    re.IGNORECASE,
)

ROOM_ONLY_RE = re.compile(
    r"^\d{1,3}[Р°-СЏС‘]?$",
    re.IGNORECASE,
)

MONTHS_RU = [
    "СЏРЅРІР°СЂСЏ", "С„РµРІСЂР°Р»СЏ", "РјР°СЂС‚Р°", "Р°РїСЂРµР»СЏ", "РјР°СЏ", "РёСЋРЅСЏ",
    "РёСЋР»СЏ", "Р°РІРіСѓСЃС‚Р°", "СЃРµРЅС‚СЏР±СЂСЏ", "РѕРєС‚СЏР±СЂСЏ", "РЅРѕСЏР±СЂСЏ", "РґРµРєР°Р±СЂСЏ",
]


# ============================================================
# DOWNLOAD
# ============================================================

def download_xlsx():
    if LOCAL_XLSX:
        print(f"LOCAL_XLSX={LOCAL_XLSX}")
        with open(LOCAL_XLSX, "rb") as f:
            return f.read()

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 Chrome/131 Safari/537.36"
        )
    }

    last_error = None

    for attempt in range(1, 4):
        try:
            print(f"Downloading XLSX (attempt {attempt}/3)...")
            response = requests.get(
                URL,
                headers=headers,
                timeout=(15, 60),
            )
            response.raise_for_status()

            content = response.content

            if not content.startswith(b"PK"):
                raise ValueError("Downloaded file is not a valid XLSX/ZIP file")

            if len(content) < 10_000:
                raise ValueError(
                    f"Downloaded XLSX is suspiciously small: {len(content)} bytes"
                )

            print(f"Downloaded: {len(content):,} bytes")
            return content

        except Exception as exc:
            last_error = exc
            print(f"Download error: {exc}")

    raise RuntimeError(f"Could not download XLSX: {last_error}")


# ============================================================
# DATE
# ============================================================

def format_date(value):
    if isinstance(value, datetime.datetime):
        value = value.date()

    return f"{value.day} {MONTHS_RU[value.month - 1]} {value.year}"


DATE_PATTERNS = [
    re.compile(r"\b\d{1,2}\s+[Рђ-РЇР°-СЏРЃС‘]+\s+\d{4}\b"),
    re.compile(r"\b\d{1,2}\.\d{1,2}\.\d{2,4}\b"),
]


def find_date(ws):
    # РЎРЅР°С‡Р°Р»Р° РёС‰РµРј РЅР°СЃС‚РѕСЏС‰СѓСЋ Excel date.
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row, 20),
        min_col=1,
        max_col=min(ws.max_column, 15),
    ):
        for cell in row:
            value = cell.value

            if isinstance(value, (datetime.datetime, datetime.date)):
                return format_date(value)

    # Р—Р°С‚РµРј РґР°С‚Сѓ РІРЅСѓС‚СЂРё С‚РµРєСЃС‚Р°.
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row, 20),
        min_col=1,
        max_col=min(ws.max_column, 15),
    ):
        for cell in row:
            if cell.value is None:
                continue

            text = str(cell.value).strip()

            for pattern in DATE_PATTERNS:
                match = pattern.search(text)
                if match:
                    return match.group(0)

    return "РќРµРёР·РІРµСЃС‚РЅР°СЏ РґР°С‚Р°"


# ============================================================
# CELL NORMALIZATION
# ============================================================

def clean_text(value):
    if value is None:
        return ""

    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    lines = []
    for line in text.split("\n"):
        line = re.sub(r"[ \t]+", " ", line).strip()
        if line:
            lines.append(line)

    return "\n".join(lines)


def extract_room(text):
    rooms = []

    def replace_room(match):
        room = match.group(1).strip()
        if room not in rooms:
            rooms.append(room)
        return ""

    text = ROOM_IN_BRACKETS_RE.sub(replace_room, text)

    return text, rooms


def is_teacher(text):
    return bool(TEACHER_RE.fullmatch(text.strip()))


def strip_subject_code(text):
    text = text.strip()

    if CODE_ONLY_RE.fullmatch(text):
        return ""

    match = CODE_WITH_TEXT_RE.match(text)
    if match:
        return match.group(2).strip()

    return text


def parse_cell(value):
    """
    РџСЂРµРѕР±СЂР°Р·СѓРµС‚ РѕРґРЅСѓ СЏС‡РµР№РєСѓ Excel РІРёРґР°:

        РћРћР”.04
        РРЅРѕСЃС‚СЂР°РЅРЅС‹Р№ СЏР·С‹Рє
        Р’Р°РЅСЏРІРёРЅР° Рћ.Рћ. (23)

    РёР»Рё:

        РћРџ.09* Р­Р»РµРєС‚СЂРѕСЂР°РґРёРѕРёР·РјРµСЂРµРЅРёСЏ
        Р“РѕСЂРёРЅ Р®.Р“. (33)

    РІ:

        {
            "subject": "...",
            "teacher": "...",
            "room": "..."
        }
    """

    text = clean_text(value)

    if not text:
        return None

    text, rooms = extract_room(text)

    lines = [line.strip() for line in text.split("\n") if line.strip()]

    # РРЅРѕРіРґР° Р°СѓРґРёС‚РѕСЂРёСЏ Р·Р°РїРёСЃР°РЅР° РѕС‚РґРµР»СЊРЅРѕР№ СЃС‚СЂРѕРєРѕР№.
    filtered_lines = []
    for line in lines:
        if ROOM_ONLY_RE.fullmatch(line):
            if line not in rooms:
                rooms.append(line)
        else:
            filtered_lines.append(line)

    lines = filtered_lines

    if not lines:
        return None

    teachers = []
    subject_parts = []

    for index, line in enumerate(lines):
        line = re.sub(r"\.\s*\.$", ".", line).strip()

        if is_teacher(line):
            if line not in teachers:
                teachers.append(line)
            continue

        # Р•СЃР»Рё РІ СЃС‚СЂРѕРєРµ РЅР°С…РѕРґРёС‚СЃСЏ РїСЂРµРїРѕРґР°РІР°С‚РµР»СЊ РІРјРµСЃС‚Рµ СЃ С‚РµРєСЃС‚РѕРј,
        # РїС‹С‚Р°РµРјСЃСЏ РѕС‚РґРµР»РёС‚СЊ РµРіРѕ.
        teacher_match = re.search(
            r"([Рђ-РЇРЃ][Р°-СЏС‘-]+(?:\s+[Рђ-РЇРЃ]\.?\s*[Рђ-РЇРЃ]\.?))\s*$",
            line,
            re.IGNORECASE,
        )

        if teacher_match:
            teacher = teacher_match.group(1).strip()
            prefix = line[:teacher_match.start()].strip()

            if teacher not in teachers:
                teachers.append(teacher)

            if prefix:
                subject_parts.append(prefix)

            continue

        subject_parts.append(line)

    # РџРµСЂРІР°СЏ СЃС‚СЂРѕРєР° РѕР±С‹С‡РЅРѕ СЃРѕРґРµСЂР¶РёС‚ РєРѕРґ РґРёСЃС†РёРїР»РёРЅС‹.
    cleaned_subject_parts = []

    for part in subject_parts:
        part = strip_subject_code(part)
        if part:
            cleaned_subject_parts.append(part)

    subject = " ".join(cleaned_subject_parts)
    subject = re.sub(r"\s+", " ", subject).strip()

    # РЇС‡РµР№РєР°, СЃРѕСЃС‚РѕСЏС‰Р°СЏ С‚РѕР»СЊРєРѕ РёР· СЃР»СѓР¶РµР±РЅРѕР№ РёРЅС„РѕСЂРјР°С†РёРё, РЅРµ СЏРІР»СЏРµС‚СЃСЏ СѓСЂРѕРєРѕРј.
    if not subject and not teachers:
        return None

    return {
        "subject": subject,
        "teacher": " / ".join(teachers),
        "room": " / ".join(rooms),
    }


# ============================================================
# MERGED CELLS
# ============================================================

def build_merged_value_map(ws):
    """
    РќРµ СЂР°Р·СЉРµРґРёРЅСЏРµС‚ merged cells.

    Р”Р»СЏ РєР°Р¶РґРѕР№ РєРѕРѕСЂРґРёРЅР°С‚С‹ merged РґРёР°РїР°Р·РѕРЅР° РІРѕР·РІСЂР°С‰Р°РµС‚ Р·РЅР°С‡РµРЅРёРµ
    РІРµСЂС…РЅРµР№ Р»РµРІРѕР№ СЏС‡РµР№РєРё. Р­С‚Рѕ Р±РµР·РѕРїР°СЃРЅРµРµ, С‡РµРј unmerge_cells(),
    РїРѕС‚РѕРјСѓ С‡С‚Рѕ РёСЃС…РѕРґРЅС‹Р№ XLSX РЅРµ РёР·РјРµРЅСЏРµС‚СЃСЏ.
    """

    merged_values = {}

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        value = ws.cell(min_row, min_col).value

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_values[(row, col)] = value

    return merged_values


def get_value(ws, merged_values, row, col):
    key = (row, col)

    if key in merged_values:
        return merged_values[key]

    return ws.cell(row=row, column=col).value


# ============================================================
# GROUPS
# ============================================================

def find_groups(ws, merged_values):
    groups = []
    group_cols = {}

    # РћСЃРЅРѕРІРЅРѕР№ РІР°СЂРёР°РЅС‚ вЂ” СЃС‚СЂРѕРєР° 2.
    for col in range(1, ws.max_column + 1):
        value = get_value(ws, merged_values, 2, col)

        if value is None:
            continue

        group = str(value).strip()

        if GROUP_RE.fullmatch(group):
            if group not in group_cols:
                groups.append(group)
                group_cols[group] = col

    # Р•СЃР»Рё СЃР°Р№С‚ РєРѕРіРґР°-РЅРёР±СѓРґСЊ РЅРµРјРЅРѕРіРѕ РёР·РјРµРЅРёС‚ С€Р°РїРєСѓ,
    # РёС‰РµРј РіСЂСѓРїРїС‹ РІ РїРµСЂРІС‹С… 10 СЃС‚СЂРѕРєР°С….
    if not groups:
        for row in range(1, min(ws.max_row, 10) + 1):
            for col in range(1, ws.max_column + 1):
                value = get_value(ws, merged_values, row, col)

                if value is None:
                    continue

                group = str(value).strip()

                if GROUP_RE.fullmatch(group):
                    if group not in group_cols:
                        groups.append(group)
                        group_cols[group] = col

    if not groups:
        raise ValueError("Groups not found in XLSX")

    return groups, group_cols


# ============================================================
# LESSON ROWS
# ============================================================

def find_lesson_rows(ws):
    rows = []

    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 1).value

        if value is None:
            continue

        if isinstance(value, bool):
            continue

        try:
            number = int(str(value).strip())
        except (ValueError, TypeError):
            continue

        if 1 <= number <= MAX_LESSON:
            rows.append((row, number))

    # РЈР±РёСЂР°РµРј РґСѓР±Р»Рё Рё РјСѓСЃРѕСЂ.
    result = []
    seen = set()

    for row, number in rows:
        if number in seen:
            continue

        seen.add(number)
        result.append((row, number))

    return result


# ============================================================
# LESSON PARSING
# ============================================================

def merge_items(items):
    """
    РћР±СЉРµРґРёРЅСЏРµС‚ РЅРµСЃРєРѕР»СЊРєРѕ Р·Р°РїРёСЃРµР№ РѕРґРЅРѕР№ РїР°СЂС‹.

    Р­С‚Рѕ РЅСѓР¶РЅРѕ РґР»СЏ "РћР±СЉРµРґРёРЅС‘РЅРЅР°СЏ/Р Р°Р·РґРµР»С‘РЅРЅР°СЏ РїР°СЂР°", РєРѕРіРґР°
    РѕРґРЅР° РіСЂСѓРїРїР° РёРјРµРµС‚ РґРІРµ Р·Р°РїРёСЃРё РІ СЃС‚СЂРѕРєР°С… start Рё start+1.
    """

    unique = []
    seen = set()

    for item in items:
        key = (
            item.get("subject", ""),
            item.get("teacher", ""),
            item.get("room", ""),
        )

        if not any(key):
            continue

        if key in seen:
            continue

        seen.add(key)
        unique.append(item)

    return unique


def make_lesson(number, items):
    items = merge_items(items)

    if not items:
        return None

    subjects = []
    teachers = []
    rooms = []

    for item in items:
        if item["subject"] and item["subject"] not in subjects:
            subjects.append(item["subject"])

        if item["teacher"] and item["teacher"] not in teachers:
            teachers.append(item["teacher"])

        if item["room"] and item["room"] not in rooms:
            rooms.append(item["room"])

    return {
        "num": str(number),

        # РЎС‚Р°СЂС‹Р№ С„РѕСЂРјР°С‚ вЂ” РѕСЃС‚Р°РІР»СЏРµРј РґР»СЏ СЃРѕРІРјРµСЃС‚РёРјРѕСЃС‚Рё СЃ С‚РµРєСѓС‰РёРј app.js.
        "subject": " / ".join(subjects),
        "teacher": " / ".join(teachers),
        "room": " / ".join(rooms),

        # РќРѕРІС‹Р№ С„РѕСЂРјР°С‚ вЂ” С‚РѕС‡РЅС‹Рµ Р·Р°РїРёСЃРё РІРЅСѓС‚СЂРё РїР°СЂС‹.
        "items": items,
    }


def parse_schedule(ws, groups, group_cols, merged_values):
    lesson_rows = find_lesson_rows(ws)

    if not lesson_rows:
        raise ValueError("Lesson rows not found in XLSX")

    print("Lesson rows:", lesson_rows)

    schedule = {group: [] for group in groups}

    for index, (start_row, number) in enumerate(lesson_rows):
        # Р’ СЌС‚РѕРј С„РѕСЂРјР°С‚Рµ СЃС‚СЂРѕРєРё start Рё start+1 СЏРІР»СЏСЋС‚СЃСЏ
        # РѕСЃРЅРѕРІРЅРѕР№ Р·Р°РїРёСЃСЊСЋ РїР°СЂС‹ Рё РІС‚РѕСЂРѕР№ Р·Р°РїРёСЃСЊСЋ (РµСЃР»Рё РµСЃС‚СЊ).
        #
        # РњС‹ РќР• С‡РёС‚Р°РµРј СЃС‚СЂРѕРєРё start+2...start+11:
        # С‚Р°Рј РЅР°С…РѕРґРёС‚СЃСЏ СЃР»СѓР¶РµР±РЅР°СЏ/СЂР°Р·Р»РѕР¶РµРЅРЅР°СЏ С‚Р°Р±Р»РёС†Р°,
        # РєРѕС‚РѕСЂР°СЏ РёРЅР°С‡Рµ СЃРѕР·РґР°С‘С‚ РґСѓР±Р»Рё Рё Р»РѕР¶РЅС‹Рµ РїСЂРµРґРјРµС‚С‹.
        candidate_rows = [start_row]

        if start_row + 1 <= ws.max_row:
            candidate_rows.append(start_row + 1)

        for group in groups:
            col = group_cols[group]
            items = []

            for row in candidate_rows:
                value = get_value(ws, merged_values, row, col)

                if value is None:
                    continue

                parsed = parse_cell(value)

                if parsed and (
                    parsed["subject"]
                    or parsed["teacher"]
                    or parsed["room"]
                ):
                    items.append(parsed)

            lesson = make_lesson(number, items)

            if lesson is not None:
                schedule[group].append(lesson)

    # РЎРѕСЂС‚РёСЂРѕРІРєР° РїРѕ РЅРѕРјРµСЂСѓ РїР°СЂС‹.
    for group in schedule:
        schedule[group].sort(key=lambda x: int(x["num"]))

    return schedule


# ============================================================
# VALIDATION
# ============================================================

def validate_result(result):
    groups = result.get("groups")

    if not isinstance(groups, list) or not groups:
        raise ValueError("Parsed result contains no groups")

    schedule = result.get("schedule")

    if not isinstance(schedule, dict):
        raise ValueError("Parsed result contains no schedule")

    lessons_count = 0

    for group in groups:
        if group not in schedule:
            raise ValueError(f"Missing schedule for group {group}")

        for lesson in schedule[group]:
            lessons_count += 1

            if not lesson.get("num"):
                raise ValueError(f"Lesson without number in group {group}")

            if not (
                lesson.get("subject")
                or lesson.get("teacher")
                or lesson.get("room")
            ):
                raise ValueError(
                    f"Empty lesson {lesson.get('num')} in group {group}"
                )

    if lessons_count == 0:
        raise ValueError("No lessons parsed from XLSX")

    print(f"Validation OK: {len(groups)} groups, {lessons_count} lessons")


# ============================================================
# ATOMIC FILE WRITE
# ============================================================

def atomic_write_json(path, data):
    directory = os.path.dirname(os.path.abspath(path)) or "."

    fd, temp_path = tempfile.mkstemp(
        prefix=".schedule_",
        suffix=".json",
        dir=directory,
        text=True,
    )

    try:
        with os.fdopen(fd, "w", encoding="utf-8") as file:
            json.dump(
                data,
                file,
                ensure_ascii=False,
                indent=2,
            )
            file.write("\n")

        os.replace(temp_path, path)

    except Exception:
        try:
            os.unlink(temp_path)
        except OSError:
            pass
        raise


def atomic_write_text(path, text):
    directory = os.path.dirname(os.path.abspath(path)) or "."

    fd, temp_path = tempfile.mkstemp(
        prefix=".schedule_",
        suffix=".tmp",
        dir=directory,
        text=True,
    )

    try:
        with os.fdopen(fd, "w", encoding="utf-8") as file:
            file.write(text)

        os.replace(temp_path, path)

    except Exception:
        try:
            os.unlink(temp_path)
        except OSError:
            pass
        raise


# ============================================================
# MAIN
# ============================================================

def main():
    try:
        content = download_xlsx()

        new_hash = hashlib.md5(content).hexdigest()

        if not LOCAL_XLSX:
            try:
                with open(HASH_FILE, "r", encoding="utf-8") as file:
                    old_hash = file.read().strip()
            except FileNotFoundError:
                old_hash = ""

            if old_hash == new_hash:
                print("No changes.")
                return

            print(
                f"Changed! "
                f"{old_hash[:8] if old_hash else 'none'} -> {new_hash[:8]}"
            )

        workbook = load_workbook(
            filename=BytesIO(content),
            data_only=True,
            read_only=False,
        )

        try:
            ws = workbook.active

            if ws.max_row < 5 or ws.max_column < 3:
                raise ValueError(
                    f"Unexpected XLSX dimensions: "
                    f"{ws.max_row}x{ws.max_column}"
                )

            merged_values = build_merged_value_map(ws)

            date = find_date(ws)
            groups, group_cols = find_groups(ws, merged_values)

            print(f"Found {len(groups)} groups")
            print(f"Date: {date}")

            schedule = parse_schedule(
                ws,
                groups,
                group_cols,
                merged_values,
            )

        finally:
            workbook.close()

        result = {
            "date": date,
            "groups": sorted(groups),
            "schedule": schedule,
        }

        validate_result(result)

        # JSON РјРµРЅСЏРµРј С‚РѕР»СЊРєРѕ РїРѕСЃР»Рµ РїРѕР»РЅРѕР№ СѓСЃРїРµС€РЅРѕР№ РїСЂРѕРІРµСЂРєРё.
        atomic_write_json(JSON_FILE, result)

        # HASH С‚РѕР¶Рµ РјРµРЅСЏРµРј С‚РѕР»СЊРєРѕ РїРѕСЃР»Рµ СѓСЃРїРµС€РЅРѕРіРѕ JSON.
        if not LOCAL_XLSX:
            atomic_write_text(HASH_FILE, new_hash + "\n")

        total = sum(len(v) for v in schedule.values())

        print(
            f"Done! "
            f"{len(groups)} groups, "
            f"{total} lessons, "
            f"date: {date}"
        )

    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
